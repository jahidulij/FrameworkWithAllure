import openpyxl
from selenium.webdriver.common.by import By


class OnlineCustomerListDDT:
    button_online_customer_xpath = "//p[contains(text(),'Online customers')]"
    headers_elements_xpath = "//*[@id='onlinecustomers-grid_wrapper']/div[1]/div/div/div[1]/div/table/thead/tr/th"
    table_info_elements_xpath = "//*[@id='onlinecustomers-grid']/tbody/tr/td"
    header_customer_info_xpath = "//th[contains(text(),'Customer info')]"
    table_ip_add_xpath = "//td[contains(text(),'172.71.98.93')]"
    path = "C:/Users/JahidulIslam/PycharmProjects/pytestSeleniumEasy2/StoreProject/data/test.xlsx"

    def __init__(self, driver):
        self.driver = driver
        self.wb = openpyxl.Workbook(self.path)
        self.sheet = self.wb.active

    def clickOnOnlineCustomers(self):
        self.driver.find_element(By.XPATH, self.button_online_customer_xpath).click()

    def getClientHeaderInfo(self):
        headers_list = []
        self.headers_info = self.driver.find_elements(By.XPATH, self.headers_elements_xpath)

        for tbl_header in self.headers_info:
            headers_list.append(tbl_header.text)
        print(headers_list)
        print(len(headers_list))
        i = 1
        while i < len(headers_list):
            # cell_num = self.sheet.cell(row=1, column=i)
            # print(cell_num)
            # print(cell_num.value)
            self.sheet.cell(row=3, column=3).value = "Hello World"
            print(headers_list[i-1])

            self.sheet.cell(row=1, column=i).value = headers_list[i-1]

            i += 1

            self.wb.save(self.path)
        # for tbl_header in self.headers_info:
        # print(tbl_header.text)

    def getClientTableInfo(self):
        self.table_content = self.driver.find_elements(By.XPATH, self.table_info_elements_xpath)
        # for tbl_header in self.table_content:
        #     print(tbl_header.text)

    def verifyHeaderContent(self):
        content = self.driver.find_element(By.XPATH, self.header_customer_info_xpath).text
        return content
